import os
import smtplib
import pandas as pd
from email.message import EmailMessage
from openpyxl import Workbook
from datetime import datetime
from playwright.sync_api import sync_playwright

CATEGORIAS = [
    "https://www.proveeduriapiaf.com.ar/categoria/carnes/",
    "https://www.proveeduriapiaf.com.ar/categoria/embutidos-rebozados/",
]

MAX_PAGINAS = 50


def extraer_precio(item):
    # Si hay oferta, el precio rebajado (rojo) esta dentro de <ins>; el tachado dentro de <del>.
    ins_loc = item.locator("ins .woocommerce-Price-amount bdi")
    if ins_loc.count() > 0:
        try:
            return ins_loc.first.text_content(timeout=2000).strip()
        except Exception:
            pass
    # Sin oferta: tomar cualquier precio (no hay <del> que confunda).
    try:
        return item.locator(".woocommerce-Price-amount bdi").first.text_content(timeout=2000).strip()
    except Exception:
        try:
            return item.locator(".woocommerce-Price-amount").first.text_content(timeout=2000).strip()
        except Exception:
            return "No encontrado"


def scrape_categoria(page, url_categoria, productos_vistos, resultados):
    for num_pagina in range(1, MAX_PAGINAS + 1):
        url = url_categoria if num_pagina == 1 else f"{url_categoria}page/{num_pagina}/"
        print(f"\n--- Pagina {num_pagina}: {url}")

        try:
            response = page.goto(url, timeout=60000)
            if response and response.status == 404:
                print("   404, fin de paginacion")
                break
            page.wait_for_timeout(2500)

            try:
                page.keyboard.press("Escape")
                page.wait_for_timeout(300)
            except Exception:
                pass

            items = page.locator("div.product-grid-item")
            count = items.count()
            print(f"   {count} productos encontrados")

            if count == 0:
                print("   Sin productos, fin de paginacion")
                break

            nuevos_en_pagina = 0
            for i in range(count):
                item = items.nth(i)
                try:
                    nombre = item.locator("h3.product-title a").text_content(timeout=3000).strip()
                except Exception:
                    nombre = ""

                precio = extraer_precio(item)

                if nombre and nombre not in productos_vistos:
                    productos_vistos.add(nombre)
                    resultados.append({
                        "Producto": nombre,
                        "Precio": precio,
                    })
                    print(f"   {nombre} | {precio}")
                    nuevos_en_pagina += 1

            if nuevos_en_pagina == 0:
                print("   Sin productos nuevos, fin de paginacion")
                break

        except Exception as e:
            print(f"   Error en pagina {num_pagina}: {e}")
            break


def scrape_piaf():
    resultados = []
    productos_vistos = set()
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        for url_categoria in CATEGORIAS:
            print(f"\n=== Categoria: {url_categoria} ===")
            scrape_categoria(page, url_categoria, productos_vistos, resultados)

        browser.close()
    return pd.DataFrame(resultados)


def enviar_mail(nombre_archivo, cantidad_productos):
    remitente = os.environ["MAIL_REMITENTE"]
    password = os.environ["MAIL_PASSWORD"]
    destinatario = os.environ["MAIL_DESTINATARIO"]

    msg = EmailMessage()
    msg["Subject"] = f"Precios Piaf - {datetime.now().strftime('%Y-%m-%d')}"
    msg["From"] = remitente
    msg["To"] = destinatario
    msg.set_content(
        f"Adjunto el listado de precios de Proveeduria Piaf.\n\n"
        f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
        f"Productos scrapeados: {cantidad_productos}\n"
    )

    with open(nombre_archivo, "rb") as f:
        contenido = f.read()
    msg.add_attachment(
        contenido,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=nombre_archivo,
    )

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(remitente, password)
        smtp.send_message(msg)

    print(f"Mail enviado a {destinatario}")


if __name__ == "__main__":
    df_resultados = scrape_piaf()

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Proveedor"
    ws["B1"] = "Proveeduria Piaf"
    ws["A2"] = "Fecha vigente"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d")

    # Encabezados de datos en fila 3
    ws["A3"] = "Producto"
    ws["B3"] = "Precio"

    # Escribir datos desde fila 4
    for idx, row in df_resultados.iterrows():
        ws[f"A{idx + 4}"] = row["Producto"]
        ws[f"B{idx + 4}"] = row["Precio"]

    nombre_archivo = f"Piaf V {datetime.now().strftime('%Y-%m-%d')}.xlsx"
    wb.save(nombre_archivo)
    print(f"\nArchivo generado: {nombre_archivo} ({len(df_resultados)} productos)")

    enviar_mail(nombre_archivo, len(df_resultados))
